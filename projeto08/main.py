import tkinter as tk
from tkinter import ttk
from openpyxl import load_workbook, Workbook 
from openpyxl.utils import get_column_letter 
from datetime import date, datetime, timedelta 
import matplotlib.pyplot as plt 
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg 
import calendar 


# funcao salvar valor
def salvar_valor():
    valor_dia = float(entry_valor.get())
    valores.append(valor_dia)
    total_valores = sum(valores)
    entry_valor.delete(0, tk.END)

    label_status.config(text="Valor salvo com Sucesso!", foreground="green")
    label_total.config(text=f"Total Ganho: R${total_valores:.2f}")

    linha = len(valores) + 1
    coluna_data = get_column_letter(1)
    coluna_valor = get_column_letter(2)
    sheet.cell(row=linha, column=1, value=date.today().strftime("%d-%m-%y"))
    sheet.cell(row=linha, column=2, value=valor_dia)


def plotar_grafico():
    global canvas

    if sheet.max_row == 0 or not any(cell.value for cell in sheet['A'][1:]):
        label_status.config(text="Nenhum dado encontrado!", foreground="red")
        return

    datas = [cell.value.date() if isinstance(cell.value, datetime) else datetime.strptime(cell.value, "%d-%m-%y").date() for cell in sheet['A'][1:]]
    valores = [cell.value for cell in sheet['B'][1:]]

    if not datas:
        label_status.config(text="Nenhuma data encontrada!", foreground="red")
        return

    dados_mensais = {}
    for cell in sheet['A'][1:]:
        if isinstance(cell.value, datetime):
            data = cell.value.date()
        else:
            data = datetime.strptime(cell.value, "%d-%m-%y").date()

        valor = sheet['B'][cell.row - 1].value

        mes_ano = data.strftime("%m-%y")

        if mes_ano in dados_mensais:
            dados_mensais[mes_ano].append(valor)
        else:
            dados_mensais[mes_ano] = [valor]
        
    fig = plt.figure(figsize=(12, 6), dpi=80)

    ax_barras = fig.add_subplot(121)
    ax_pie = fig.add_subplot(122)

    barras = ax_barras.bar(range(len(dados_mensais)), [sum(valores) for valores in dados_mensais.values()])
    
    for i, barra in enumerate(barras):
        altura = barra.get_height()
        ax_barras.text(barra.get_x() + barra.get_width() / 2, altura, f'R${altura:.2f}', ha="center", va="bottom")

    nomes_meses = []
    for mes_ano in dados_mensais.keys():
        mes, ano = mes_ano.split('-')
        nome_mes = calendar.month_name[int(mes)]
        nomes_meses.append(f'{nome_mes}-{ano}')

    ax_barras.set_xticks(range(len(dados_mensais)))
    ax_barras.set_xticklabels(nomes_meses, ha='right')

    ax_barras.spines['top'].set_visible(False)
    ax_barras.spines['right'].set_visible(False)
    ax_barras.spines['bottom'].set_visible(False)
    ax_barras.spines['left'].set_visible(False)

    ax_barras.set_title("Economia Por mês", fontweight="bold")

    ax_barras.title.set_position([.5, 8.05])
    ax_barras.set_xlabel('Mês')
    ax_barras.set_ylabel('Valor Ganho')

    datas = [cell.value.date() if isinstance(cell.value, datetime) else datetime.strptime(cell.value, "%d-%m-%y").date() for cell in sheet['A'][1:]]

    data_inicial = min(datas)
    data_final = max(datas)

    diferenca = (data_final - data_inicial).days
    semanas = diferenca // 7

    labels = [f'{i+1}° Semana' for i in range(semanas)]
    valores_semana = []
    for i in range(semanas):
        inicio_semana = data_inicial + timedelta(days=i * 7)
        fim_semana = inicio_semana + timedelta(days=6)
        valores_semana.append(sum([valor for data, valor in zip(datas, valores) if inicio_semana <= data <= fim_semana]))

    ax_pie.pie(valores_semana, labels=None, startangle=90, labeldistance=0.7)

    ax_pie.axis('equal')

    labels_valores = [f'{label}\nR${valor:.2f}' for label, valor in zip(labels, valores_semana)]

    ax_pie.legend(labels=labels_valores, loc='center left', bbox_to_anchor=(0.75, 0.5), ncol=1)

    ax_pie.set_title('Ganhos Por Semana')

    canvas = FigureCanvasTkAgg(fig, master=janela)
    canvas.draw()
    canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True)



janela = tk.Tk()
janela.title("Aplicatico de Poupança")
janela.geometry("700x500")
janela.configure(bg="#252525")

style = ttk.Style()
style.theme_use("clam")
style.configure("TLabel", background="#252525", foreground="#FFFFFF", font=("Arial", 12))
style.configure("TEntry", fieldbackground="#FFFFFF", font=("Arial", 12))
style.configure("TButton", background="#4CAF50", foreground="#FFFFFF", font=("Arial", 12))

label_instrucao = ttk.Label(janela, text="Insira o valor diario")
label_status = ttk.Label(janela, text="", foreground="red")
label_total = ttk.Label(janela, text="100", font=("Arial", 14, "bold"))
entry_valor = ttk.Entry(janela)
button_salvar = ttk.Button(janela, text="Salvar", command=salvar_valor)

label_instrucao.pack(pady=5)
entry_valor.pack(pady=5)
button_salvar.pack(pady=10)
label_status.pack()
label_total.pack(pady=10)

try:
    Workbook = load_workbook("valores_diarios.xlsx")
except FileNotFoundError:
    Workbook = Workbook()

sheet = Workbook.active

if sheet.max_row == 0:
    sheet.cell(row=1, column=1, value="Data")
    sheet.cell(row=1, column=2, value="Valor diário")
    
valores = [cell.value for cell in sheet['B'][1:]]

label_total.config(text=f'Total Ganho R${sum(valores):.2f}')

plotar_grafico()

janela.mainloop()

Workbook.save("valores_diarios.xlsx")

